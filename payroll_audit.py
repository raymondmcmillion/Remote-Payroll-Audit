# =============================================================================
# PAYROLL AUDIT — Remote → Gusto
# =============================================================================
#
# PURPOSE
# -------
# Compares Remote payroll input against Gusto payroll output to flag
# discrepancies before payroll is finalized. Produces a multi-sheet Excel
# workbook. Run once per pay period by updating the CONFIG section below.
#
# INPUT FILES (update CONFIG for each run)
# -----------------------------------------
# 1. Remote Input  : Excel workbook with multiple sheets (see INPUT SHEETS below)
# 2. Gusto Output  : CSV with a single clean header row
# 3. Mapping File  : CSV linking Remote employee names to Gusto names + dept
# 4. name_overrides.csv : persistent manual name pairings (same folder as script)
#
# INPUT SHEETS USED
# -----------------
# - "Payroll summary" : base salary, annual salary, join/term dates, status,
#                       Total Allowance, Total Stipend, Total Expenses, Employment ID
# - "Pay Items"       : SOURCE OF TRUTH for all pay items (allowances, WFH stipend,
#                       expenses, retro, overtime, other earnings). ALWAYS use this
#                       sheet to compare pay items against Gusto output columns.
#                       Do NOT rely solely on Payroll summary for pay item amounts.
#
# PAY PERIOD RULES
# ----------------
# - Semi-monthly: 24 pay periods/year
# - Full period proration : Annual ÷ 24
# - Partial period (new hire or term): Annual ÷ 260 × working days
# - Working days = Monday–Friday only (holidays NOT excluded)
# - July 16–31, 2026 = 12 working days
#
# NAME MATCHING (6-level fallback, in order)
# ------------------------------------------
# 0. name_overrides.csv   — manual overrides for confirmed cross-system pairings
#    (e.g. Josh Hall = Hall, Joshua; Mitchell Gray = Gray III, Robert)
#    ADD NEW PAIRS HERE when a match fails; do not delete existing entries.
# 1. Exact remote name match against mapping
# 2. First + last word match against mapping
# 3. col_b first-word match
# 4. Suffix fallback (last-word, first-word) — handles "Gray III", "Kempton Jr"
# 5. Word-subset fallback — handles compound names
# 6. Direct output lookup (last+first word of Remote name against Gusto output)
#    — catches new hires not yet in mapping
# matched_out_keys guards prevent any Gusto row from being claimed twice.
#
# LOA DETECTION
# -------------
# Mapping file col 4 = "LOA" string → employee is on Leave of Absence.
# LOA employees are NOT expected in Gusto output; they go in Missing → LOA section.
#
# DEPARTMENT SLUG (dept_slug function)
# -------------------------------------
# Strips letter-range suffix from department name.
# Example: "RE (01) A-Bl" → "RE (01)", "RI (2) G-Q" → "RI (2)"
# "LOA", "Terms" are left unchanged.
# Always display dept_slug(dept) on every sheet, never the raw dept string.
#
# TERMS DEPARTMENT
# ----------------
# Gusto output has a "Terms" department for terminated employees.
# Department is pulled from out_row[2] automatically — no special handling needed.
#
# EMPLOYMENT ID
# -------------
# - Always include Employment ID as the second column on EVERY sheet.
# - Source for matched/input rows: ir[IN_EMP_ID] (col 0 of input file)
# - Source for output-only rows: gusto_to_emp_id lookup (built from mapping col 2)
#
# OUTPUT SHEETS PRODUCED
# -----------------------
# 1.  Summary              — counts for every sheet
# 2.  Base Salary Variances— matched pairs where period gross ≠ output base
#                            (Regular + PTO). Color: Red=Needs Review,
#                            Yellow=Term, Green=New Hire, Blue=LOA/Term dept.
# 3.  Term Sheet           — terminated employees; proration verified
# 4.  New Hire Sheet       — new hires in period; proration verified
# 5.  Pay Item Flags       — ALL pay items from Pay Items tab vs Gusto output columns.
#                            Also flags unmatched employees with pay items (output=0).
#                            Status colors: Green=Match, Red=Mismatch,
#                            Orange=Not in Gusto, Gray=No Gusto column.
# 6.  Zero or Negative     — Gusto output rows where Regular ≤ 0 or Net Pay ≤ 0
# 7.  Hours Over Threshold — matched pairs where Regular + PTO hours > 86.67
#                            (no PTO payouts counted)
# 8.  Missing Employees    — three sections:
#                            • Should be paid – not in Gusto (red header)
#                            • On LOA – not expected in Gusto (gray header)
#                            • In Gusto but not in Remote input (orange header)
# 9.  Archived Employees   — Remote employees with status = "archived"
# 10. Retro                — from Pay Items tab "Retro" rows vs Gusto Retro column.
#                            ALWAYS pull retros from Pay Items tab, not Payroll summary.
# 11. Gross Up             — Allowances + Stipends only; Payroll Summary vs Gusto.
#                            Shows ONLY mismatches and "Not in Gusto" rows.
#                            Includes archived employees with allowances if diff exists.
#
# PAY ITEM FLAGS — DETAILED RULES
# ---------------------------------
# Source of truth: "Pay Items" tab (not Payroll summary).
# Pay Items tab → Gusto output column mapping:
#   allowance      → Gusto "Allowance"
#   wfh stipend    → Gusto "WFH Stipend"
#   retro          → Gusto "Retro"
#   expense types  → Gusto "Reimbursements"
#     (expense non-taxable, expense taxable, mileage non-taxable, per diem non-taxable)
#   other earnings → Gusto "Other"
#   overtime       → Gusto "Overtime (Amount)" (flagged as "No Gusto column" if absent)
# Skipped types (handled elsewhere): regular salary, paid time off,
#   unpaid time off, other leave, correction.
#
# ALLOWANCE / STIPEND RULE
# ------------------------
# Pay Item Flags  : uses Pay Items tab amounts (what was sent to Gusto)
# Gross Up sheet  : uses Payroll Summary Total Allowance + Total Stipend (contractual)
# Both are needed. Pay Item Flags catches processing errors;
# Gross Up catches FX/proration/gross-up differences.
#
# EXPENSE RULE
# ------------
# Always use Pay Items tab as source for expenses.
# Payroll Summary "Total Expenses" can be $0 even when an expense exists
# (e.g. Sharon Jennrich — $764 showed in Pay Items but not Payroll Summary).
# Pay Items tab is authoritative.
#
# RETRO RULE
# ----------
# Retro is on the "Pay Items" tab, NOT on the Payroll summary tab.
# Always read retros from Pay Items. Compare against Gusto "Retro" column.
# A dedicated "Retro" sheet shows all retros with input vs output and match status.
#
# HOURS THRESHOLD
# ---------------
# Flag when Regular Hours + PTO Hours > 86.67.
# Do NOT include PTO payouts (Outstanding Paid Time Off) in this total.
#
# VARIANCE THRESHOLD
# ------------------
# $1.00 — differences smaller than this are ignored on all sheets.
#
# name_overrides.csv FORMAT
# -------------------------
# Remote Name, Gusto Last, Gusto First, Notes
# Add a row whenever a name pairing is manually confirmed.
# Never delete existing rows — they prevent future false mismatches.
#
# =============================================================================

import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date, timedelta

# ─── CONFIGURATION ───────────────────────────────────────────────────────────
PERIOD_START = date(2026, 7, 16)
PERIOD_END   = date(2026, 7, 31)
PAY_PERIODS_PER_YEAR = 24
ANNUAL_WORKING_DAYS  = 260
VARIANCE_THRESHOLD   = 1.00  # dollars

INPUT_PATH   = '/sessions/ecstatic-nice-johnson/mnt/uploads/2 Pre Input.xlsx'
OUTPUT_PATH  = '/sessions/ecstatic-nice-johnson/mnt/uploads/2 Pre Output.csv'
MAPPING_PATH = '/sessions/ecstatic-nice-johnson/mnt/uploads/2 Mapping File.csv'
RESULT_PATH  = '/sessions/ecstatic-nice-johnson/mnt/outputs/Payroll_Audit_July31_2026.xlsx'

# ─── WORKING DAYS ────────────────────────────────────────────────────────────
def build_working_days(start, end):
    days, d = [], start
    while d <= end:
        if d.weekday() < 5:
            days.append(d)
        d += timedelta(days=1)
    return days

PERIOD_WD = build_working_days(PERIOD_START, PERIOD_END)
TOTAL_WD   = len(PERIOD_WD)   # 12

def working_days_between(start, end):
    s = max(start, PERIOD_START)
    e = min(end,   PERIOD_END)
    if s > e:
        return 0
    return sum(1 for d in PERIOD_WD if s <= d <= e)

def parse_date(s):
    if not s or not str(s).strip():
        return None
    s = str(s).strip()
    for fmt in ['%Y-%m-%d', '%m/%d/%y', '%m/%d/%Y']:
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None

def parse_amount(s):
    if not s or not str(s).strip():
        return 0.0
    try:
        return float(str(s).strip().replace(',', '').replace('$', ''))
    except Exception:
        return 0.0

# ─── READ INPUT (XLSX — "Payroll summary" sheet) ──────────────────────────────
_wb = openpyxl.load_workbook(INPUT_PATH, data_only=True)
_ws = _wb['Payroll summary']
_rows = list(_ws.iter_rows(values_only=True))
in_headers = [str(c) if c is not None else '' for c in _rows[0]]
in_rows = [
    [str(c) if c is not None else '' for c in r]
    for r in _rows[1:] if any(c is not None and str(c).strip() for c in r)
]

# ─── READ PAY ITEMS SHEET ────────────────────────────────────────────────────
# Source of truth for ALL pay items — compared against Gusto output columns
_pi_ws   = _wb['Pay Items']
_pi_rows = list(_pi_ws.iter_rows(values_only=True))
_pi_hdrs = [str(c) if c is not None else '' for c in _pi_rows[0]]
_PI = {h: i for i, h in enumerate(_pi_hdrs)}

# Expense-type pay items that roll up into Gusto Reimbursements
EXPENSE_PI_TYPES = {'expense non-taxable', 'expense taxable',
                    'mileage non-taxable', 'per diem non-taxable'}

# Skip types that are already covered by Base Salary / Hours sheets
SKIP_PI_TYPES = {'regular salary', 'paid time off', 'unpaid time off',
                 'other leave', 'correction'}

# pi_totals[eid][canonical_type] = total amount
# canonical types: 'allowance', 'wfh stipend', 'retro', 'expense', 'other earnings', 'overtime'
from collections import defaultdict
pi_totals   = defaultdict(lambda: defaultdict(float))
pi_notes    = defaultdict(dict)   # [eid][canonical_type] = note
pi_emp_name = {}                  # eid -> name

# retro still needs note for Retro sheet
retro_by_emp_id = {}

for r in _pi_rows[1:]:
    row     = [str(c) if c is not None else '' for c in r]
    eid     = row[_PI['Employment ID']].strip()
    name    = row[_PI['Employee Name']].strip()
    item_lc = row[_PI['Pay Item Name']].strip().lower()
    amt     = parse_amount(row[_PI['Value']])
    note    = row[_PI['Note']].strip()
    if not eid or item_lc in SKIP_PI_TYPES:
        continue
    pi_emp_name[eid] = name

    if item_lc in EXPENSE_PI_TYPES:
        canon = 'expense'
    elif item_lc == 'overtime (amount)':
        canon = 'overtime'
    else:
        canon = item_lc   # 'allowance', 'wfh stipend', 'retro', 'other earnings'

    pi_totals[eid][canon] += amt
    if note and canon not in pi_notes[eid]:
        pi_notes[eid][canon] = note

    if item_lc == 'retro':
        retro_by_emp_id[eid] = {'name': name, 'amount': amt, 'note': note}

# For backward compat (Retro sheet uses these)
def pi_get(eid, canon):
    return pi_totals.get(eid, {}).get(canon, 0.0)

# ─── READ OUTPUT (CSV — clean single header) ──────────────────────────────────
with open(OUTPUT_PATH, 'r', encoding='utf-8-sig') as f:
    raw = list(csv.reader(f))

out_headers = raw[0]
out_rows = []
for r in raw[1:]:
    if not r or not r[0].strip():
        continue
    out_rows.append(r)

# ─── READ MAPPING ─────────────────────────────────────────────────────────────
with open(MAPPING_PATH, 'r', encoding='utf-8-sig') as f:
    reader = csv.reader(f)
    map_headers = next(reader)
    map_rows = list(reader)

# ─── NAME OVERRIDES ───────────────────────────────────────────────────────────
import os
OVERRIDES_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'name_overrides.csv')
name_overrides = {}
if os.path.exists(OVERRIDES_PATH):
    with open(OVERRIDES_PATH, 'r', encoding='utf-8-sig') as f:
        _or = csv.DictReader(f)
        for _row in _or:
            _rn = _row.get('Remote Name', '').strip()
            _gl = _row.get('Gusto Last',  '').strip()
            _gf = _row.get('Gusto First', '').strip()
            if _rn and _gl and _gf:
                name_overrides[_rn.lower()] = (_gl, _gf)

# ─── BUILD LOOKUPS ────────────────────────────────────────────────────────────
import unicodedata, re

SUFFIXES = {'jr', 'jr.', 'sr', 'sr.', 'ii', 'iii', 'iv', 'v', 'vi'}

def normalize(name):
    nfkd  = unicodedata.normalize('NFKD', name)
    ascii_ = nfkd.encode('ascii', 'ignore').decode('ascii')
    return ascii_.lower().strip()

def strip_suffixes(parts):
    while parts and parts[-1].lower().rstrip('.') in SUFFIXES:
        parts = parts[:-1]
    return parts

def first_last(name):
    parts = strip_suffixes(normalize(name).split())
    if len(parts) >= 2:
        return (parts[0], parts[-1])
    return None

def clean_col_b(name):
    return re.sub(r'\b(TERM|LOA|INACTIVE)\b', '', name, flags=re.IGNORECASE).strip()

def dept_slug(dept):
    """Shorten 'RE (01) A-Bl' → 'RE (01)', 'RI (2) G-Q' → 'RI (2)', 'LOA'/'Terms' unchanged."""
    return re.sub(r'\s+[A-Z][a-zA-Z]?-[A-Z][a-zA-Z]*$', '', str(dept)).strip()

remote_to_gusto    = {}
firstlast_to_gusto = {}
remote_to_loa      = {}

for row in map_rows:
    rname_raw = row[6].strip() if len(row) > 6 else ''
    if rname_raw:
        col4 = row[4].strip() if len(row) > 4 else ''
        if col4.upper() == 'LOA':
            remote_to_loa[normalize(rname_raw)] = True

    if len(row) > 6 and row[5].strip():
        gfull = row[5].strip()
        dept  = row[13].strip() if len(row) > 13 else ''
        parts = gfull.split(',', 1)
        if len(parts) != 2:
            continue
        col_b   = clean_col_b(row[1].strip()) if len(row) > 1 else ''
        b_norm  = normalize(col_b)
        b_parts = b_norm.split()
        col_b_last  = b_parts[-1]              if len(b_parts) >= 1 else ''
        col_b_first = ' '.join(b_parts[:-1])   if len(b_parts) >= 2 else ''

        info = {
            'gusto_last':  col_b_last,
            'gusto_first': col_b_first,
            'department':  dept,
        }

        rname = row[6].strip() if len(row) > 6 else ''
        if rname:
            remote_to_gusto[normalize(rname)] = info
            fl = first_last(rname)
            if fl and fl not in firstlast_to_gusto:
                firstlast_to_gusto[fl] = info

        if col_b:
            key_b = b_norm
            if key_b not in remote_to_gusto:
                remote_to_gusto[key_b] = info
            fl_b = first_last(col_b)
            if fl_b and fl_b not in firstlast_to_gusto:
                firstlast_to_gusto[fl_b] = info

colb_wordset_index = []
for row in map_rows:
    if len(row) > 1 and row[1].strip():
        col_b_clean = clean_col_b(row[1].strip())
        if not col_b_clean:
            continue
        dept  = row[13].strip() if len(row) > 13 else ''
        b_norm_parts = normalize(col_b_clean).split()
        if len(b_norm_parts) < 2:
            continue
        col_b_last  = b_norm_parts[-1]
        col_b_first = ' '.join(b_norm_parts[:-1])
        info_ws = {
            'gusto_last':  col_b_last,
            'gusto_first': col_b_first,
            'department':  dept,
        }
        colb_wordset_index.append((frozenset(b_norm_parts), info_ws))

mapping_remote_names = set(remote_to_gusto.keys())

out_lookup        = {}
out_lookup_suffix = {}
for row in out_rows:
    last  = normalize(row[0])
    first = normalize(row[1])
    key   = (last, first)
    out_lookup[key] = row
    last_word  = last.split()[-1]  if last  else ''
    first_word = first.split()[0]  if first else ''
    suf_key = (last_word, first_word)
    if suf_key not in out_lookup_suffix:
        out_lookup_suffix[suf_key] = row

# Output column indices
oc = {h: i for i, h in enumerate(out_headers)}
def oi(name): return oc.get(name, -1)

REG_AMT_IDX     = oi('Regular (Amount)')
REG_HRS_IDX     = oi('Regular (Hours)')
TO_AMT_IDX      = oi('Time Off (Amount)')
TO_HRS_IDX      = oi('Time Off (Hours)')
PTO_OUT_HRS_IDX = oi('Outstanding Paid Time Off (Hours)')
PTO_OUT_AMT_IDX = oi('Outstanding Paid Time Off (Amount)')
NET_IDX         = oi('Net Pay')
REIMB_IDX       = oi('Reimbursements')
ALLOW_OUT_IDX   = oi('Allowance')
WFH_OUT_IDX     = oi('WFH Stipend')
RETRO_IDX       = oi('Retro')
OTHER_OUT_IDX   = oi('Other')
OT_OUT_IDX      = oi('Overtime (Amount)')   # may be -1 if not in this output
HOURS_THRESHOLD = 86.67

# Pay Item → Gusto output column mapping
# (canonical_type, label, output_idx)
PAYITEM_GUSTO_MAP = [
    ('allowance',      'Allowance',       ALLOW_OUT_IDX),
    ('wfh stipend',    'WFH Stipend',     WFH_OUT_IDX),
    ('retro',          'Retro',           RETRO_IDX),
    ('expense',        'Expense/Reimb',   REIMB_IDX),
    ('other earnings', 'Other Earnings',  OTHER_OUT_IDX),
    ('overtime',       'Overtime',        OT_OUT_IDX),
]

def oval(row, idx):
    if idx < 0 or idx >= len(row): return 0.0
    return parse_amount(row[idx])

def ostr(row, idx):
    if idx < 0 or idx >= len(row): return ''
    return str(row[idx]).strip()

# Input column indices
IN = {h: i for i, h in enumerate(in_headers)}
IN_EMP_ID    = IN.get('Employment ID', 0)
IN_NAME      = IN['Employee Name']
IN_STATUS    = IN['Employee Status']
IN_ANNUAL    = IN['Annual Salary']
IN_PERIOD    = IN['Period Gross Salary']
IN_JOIN      = IN['Joining Date']
IN_TERM      = IN['Termination Date']
IN_ALLOW     = IN.get('Total Allowance', -1)
IN_ALLOW_DESC= IN.get('Allowance Description', -1)
IN_STIP      = IN.get('Total Stipend', -1)
IN_STIP_DESC = IN.get('Stipend Description', -1)
IN_OT        = IN.get('Total Overtime', -1)
IN_OT_DESC   = IN.get('Overtime Description', -1)
IN_BONUS     = IN.get('Total Bonus', -1)
IN_BONUS_DESC= IN.get('Bonus Description', -1)
IN_COMM      = IN.get('Total Commission', -1)
IN_COMM_DESC = IN.get('Commission Description', -1)
IN_EXP       = IN.get('Total Expenses', -1)
IN_EXP_DESC  = IN.get('Expenses Description', -1)
IN_OTHER     = IN.get('Total other Incentives', -1)
IN_OTHER_DESC= IN.get('Other Incentives Description', -1)

def get_col(ir, idx):
    if idx < 0 or idx >= len(ir): return ''
    return str(ir[idx]).strip()

def get_amount(ir, idx):
    return parse_amount(get_col(ir, idx))

def get_desc(ir, idx):
    return get_col(ir, idx)

# ─── MATCH EMPLOYEES ─────────────────────────────────────────────────────────
matched          = []
input_only       = []
matched_out_keys = set()

for ir in in_rows:
    if not ir or not ir[IN_NAME].strip():
        continue
    rname      = ir[IN_NAME].strip()
    rname_norm = normalize(rname)

    # 0) Manual override
    out_row = None
    dept    = ''
    if rname.lower() in name_overrides:
        _gl, _gf = name_overrides[rname.lower()]
        _key = (normalize(_gl), normalize(_gf))
        if _key in out_lookup:
            out_row = out_lookup[_key]
            dept    = out_row[2]
            matched_out_keys.add(_key)
            matched.append((ir, out_row, dept))
            continue

    # 1) Exact remote name match
    ginfo      = remote_to_gusto.get(rname_norm)
    match_type = 'exact'

    # 2) First + last fallback
    if not ginfo:
        fl = first_last(rname)
        if fl:
            ginfo      = firstlast_to_gusto.get(fl)
            match_type = 'first+last fallback'

    if ginfo:
        key = (normalize(ginfo['gusto_last']), normalize(ginfo['gusto_first']))
        if key in out_lookup:
            out_row = out_lookup[key]
            dept = out_row[2]
            matched_out_keys.add(key)
        else:
            first_word = normalize(ginfo['gusto_first']).split()[0] if ginfo['gusto_first'] else ''
            key2 = (normalize(ginfo['gusto_last']), first_word)
            if key2 in out_lookup:
                out_row = out_lookup[key2]
                dept = out_row[2]
                matched_out_keys.add(key2)

        if not out_row and ginfo.get('col_b_first'):
            key3 = (normalize(ginfo['gusto_last']), ginfo['col_b_first'])
            if key3 in out_lookup:
                out_row = out_lookup[key3]
                dept = out_row[2]
                matched_out_keys.add(key3)

        # 4) Suffix fallback
        if not out_row and ginfo['gusto_last'] and ginfo['gusto_first']:
            g_last_word  = normalize(ginfo['gusto_last']).split()[-1]
            g_first_word = normalize(ginfo['gusto_first']).split()[0] if ginfo['gusto_first'] else ''
            key4 = (g_last_word, g_first_word)
            if key4 in out_lookup_suffix:
                candidate = out_lookup_suffix[key4]
                cand_key  = (candidate[0].strip().lower(), candidate[1].strip().lower())
                if cand_key not in matched_out_keys:
                    out_row = candidate
                    dept = out_row[2]
                    matched_out_keys.add(cand_key)

        # 5) Word-subset fallback
        if not out_row:
            input_words = frozenset(normalize(rname).split())
            for colb_words, ws_info in colb_wordset_index:
                if colb_words <= input_words or input_words <= colb_words:
                    k = (ws_info['gusto_last'], ws_info['gusto_first'])
                    if k in out_lookup and k not in matched_out_keys:
                        out_row = out_lookup[k]
                        dept = out_row[2]
                        matched_out_keys.add(k)
                        ginfo = ws_info
                        break
                    k_suf = (ws_info['gusto_last'], ws_info['gusto_first'].split()[0] if ws_info['gusto_first'] else '')
                    if k_suf in out_lookup_suffix:
                        candidate = out_lookup_suffix[k_suf]
                        cand_key  = (candidate[0].strip().lower(), candidate[1].strip().lower())
                        if cand_key not in matched_out_keys:
                            out_row = candidate
                            dept = out_row[2]
                            matched_out_keys.add(cand_key)
                            ginfo = ws_info
                            break

        if not out_row:
            dept = ginfo['department'] if ginfo else ''

    # 6) Direct output fallback for employees not in mapping (new hires, etc.)
    #    Tries last-word + first-word of Remote name directly against Gusto output.
    #    Only activates when mapping lookup found nothing (ginfo is None).
    if not out_row and not ginfo:
        parts = strip_suffixes(normalize(rname).split())
        if len(parts) >= 2:
            last_w  = parts[-1]
            first_w = parts[0]
            direct_key = (last_w, first_w)
            if direct_key in out_lookup and direct_key not in matched_out_keys:
                out_row = out_lookup[direct_key]
                dept = out_row[2]
                matched_out_keys.add(direct_key)
            elif direct_key in out_lookup_suffix:
                candidate = out_lookup_suffix[direct_key]
                cand_key  = (candidate[0].strip().lower(), candidate[1].strip().lower())
                if cand_key not in matched_out_keys:
                    out_row = candidate
                    dept = out_row[2]
                    matched_out_keys.add(cand_key)

    if out_row:
        matched.append((ir, out_row, dept))
    else:
        is_loa = remote_to_loa.get(rname_norm, False)
        if is_loa:
            miss_reason = 'LOA'
        elif rname_norm not in mapping_remote_names and first_last(rname) not in {first_last(k) for k in mapping_remote_names}:
            miss_reason = 'Not in mapping file'
        elif ginfo and not out_row:
            miss_reason = 'Not in Gusto this period – verify'
        elif any(c for c in rname if ord(c) > 127):
            miss_reason = 'Special characters – verify mapping'
        else:
            miss_reason = 'Name variant – update mapping file'
        input_only.append((ir, dept, miss_reason))

output_only = [row for key, row in out_lookup.items()
               if key not in matched_out_keys]

# Employment ID lookup for Gusto-only rows
gusto_to_emp_id = {}
for _mr in map_rows:
    _eid  = _mr[2].strip() if len(_mr) > 2 else ''
    _colb = clean_col_b(_mr[1].strip()) if len(_mr) > 1 else ''
    if _eid and _colb:
        _bp = normalize(_colb).split()
        if len(_bp) >= 2:
            gusto_to_emp_id[(_bp[-1], _bp[0])] = _eid

def get_gusto_emp_id(out_row):
    last  = normalize(out_row[0]).split()
    first = normalize(out_row[1]).split()
    return gusto_to_emp_id.get((last[-1] if last else '', first[0] if first else ''), '')

# ─── STYLE HELPERS ───────────────────────────────────────────────────────────
H_FILL    = PatternFill('solid', fgColor='1F4E79')
H_FONT    = Font(color='FFFFFF', bold=True, size=10)
ALT       = PatternFill('solid', fgColor='D9E1F2')
WARN      = PatternFill('solid', fgColor='FFE699')
RED       = PatternFill('solid', fgColor='FF9999')
GREEN     = PatternFill('solid', fgColor='C6EFCE')
BLUE      = PatternFill('solid', fgColor='BDD7EE')
ORANGE    = PatternFill('solid', fgColor='F4B942')
DARK_FONT = Font(bold=True, size=10)

def header_row(ws, titles):
    ws.append(titles)
    for cell in ws[ws.max_row]:
        cell.font      = H_FONT
        cell.fill      = H_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')
    ws.row_dimensions[ws.max_row].height = 28

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze(ws): ws.freeze_panes = 'A2'

CURR = '$#,##0.00'

def fmt_currency(ws, col_letters, start_row=2):
    for cl in col_letters:
        for cell in ws[cl]:
            if cell.row >= start_row:
                cell.number_format = CURR

# ─── WORKBOOK ────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ─── SHEET 1: BASE SALARY VARIANCES ─────────────────────────────────────────
ws1 = wb.active
ws1.title = 'Base Salary Variances'

header_row(ws1, [
    'Employee Name', 'Emp ID', 'Department', 'Annual Salary',
    'Input Period Gross', 'Output Regular', 'Output PTO Taken',
    'Output Base Total', 'Variance ($)', 'Reason', 'Hire Date', 'Term Date'
])
# Note: No Paid Holidays column in July 16-31 output (no holidays in this period)
freeze(ws1)

for ir, or_, dept in matched:
    annual   = parse_amount(ir[IN_ANNUAL])
    period   = parse_amount(ir[IN_PERIOD])
    reg      = oval(or_, REG_AMT_IDX)
    to_      = oval(or_, TO_AMT_IDX)
    out_base = round(reg + to_, 2)
    variance = round(period - out_base, 2)

    if abs(variance) < VARIANCE_THRESHOLD:
        continue

    term_dt = parse_date(ir[IN_TERM])
    join_dt = parse_date(ir[IN_JOIN])

    if term_dt and PERIOD_START <= term_dt <= PERIOD_END:
        reason = 'Term'
        rfill  = WARN
    elif join_dt and PERIOD_START <= join_dt <= PERIOD_END:
        reason = 'New Hire'
        rfill  = GREEN
    elif dept.strip().upper() in ('LOA', 'TERMS', 'TERM'):
        reason = 'LOA / Term'
        rfill  = BLUE
    else:
        reason = 'Needs Review'
        rfill  = RED

    ws1.append([
        ir[IN_NAME], ir[IN_EMP_ID], dept_slug(dept), annual, period,
        reg, to_, out_base, variance, reason,
        str(ir[IN_JOIN]).strip() if ir[IN_JOIN] else '',
        str(ir[IN_TERM]).strip() if ir[IN_TERM] else ''
    ])
    r = ws1.max_row
    ws1.cell(r, 10).fill = rfill
    if r % 2 == 0:
        for c in range(1, 10):
            if ws1.cell(r, c).fill.fgColor.rgb == '00000000':
                ws1.cell(r, c).fill = ALT

fmt_currency(ws1, ['D', 'E', 'F', 'G', 'H', 'I'])
set_col_widths(ws1, [32, 10, 18, 14, 16, 15, 15, 16, 13, 14, 13, 13])

# ─── SHEET 2: TERM SHEET ─────────────────────────────────────────────────────
ws2 = wb.create_sheet('Term Sheet')

header_row(ws2, [
    'Employee Name', 'Emp ID', 'Department', 'Annual Salary',
    'Term Date', 'Working Days in Period', 'Days Worked',
    'Expected Prorated Pay', 'Output Regular', 'Output PTO Taken',
    'Output Base Total', 'Proration Match?', 'PTO Payout Hours'
])
freeze(ws2)

for ir, or_, dept in matched:
    term_dt = parse_date(ir[IN_TERM])
    if not term_dt or term_dt > PERIOD_END:
        continue

    annual      = parse_amount(ir[IN_ANNUAL])
    days_worked = working_days_between(PERIOD_START, term_dt)
    if days_worked == TOTAL_WD:
        expected_pro = round(annual / PAY_PERIODS_PER_YEAR, 2)
    else:
        expected_pro = round((annual / ANNUAL_WORKING_DAYS) * days_worked, 2)

    reg      = oval(or_, REG_AMT_IDX)
    to_      = oval(or_, TO_AMT_IDX)
    out_base = round(reg + to_, 2)
    pto_hrs  = ostr(or_, PTO_OUT_HRS_IDX)

    if abs(expected_pro - out_base) < VARIANCE_THRESHOLD:
        match_str = 'Match'
        mfill = GREEN
    else:
        match_str = 'Needs Review'
        mfill = RED

    ws2.append([
        ir[IN_NAME], ir[IN_EMP_ID], dept_slug(dept), annual,
        term_dt.strftime('%m/%d/%Y'), TOTAL_WD, days_worked,
        expected_pro, reg, to_, out_base,
        match_str, pto_hrs if pto_hrs else '—'
    ])
    r = ws2.max_row
    ws2.cell(r, 12).fill = mfill
    if r % 2 == 0:
        for c in [1,2,3,4,5,6,7,8,9,10,11,13]:
            ws2.cell(r, c).fill = ALT

fmt_currency(ws2, ['D', 'H', 'I', 'J', 'K'])
set_col_widths(ws2, [32, 10, 18, 14, 13, 14, 12, 20, 15, 15, 16, 14, 16])

# ─── SHEET 3: NEW HIRE SHEET ─────────────────────────────────────────────────
ws3 = wb.create_sheet('New Hire Sheet')

header_row(ws3, [
    'Employee Name', 'Emp ID', 'Department', 'Annual Salary',
    'Hire Date', 'Working Days in Period', 'Days Worked',
    'Expected Prorated Pay', 'Output Regular', 'Output PTO Taken',
    'Output Base Total', 'Match?'
])
freeze(ws3)

for ir, or_, dept in matched:
    join_dt = parse_date(ir[IN_JOIN])
    if not join_dt or not (PERIOD_START <= join_dt <= PERIOD_END):
        continue

    annual      = parse_amount(ir[IN_ANNUAL])
    days_worked = working_days_between(join_dt, PERIOD_END)
    if days_worked == TOTAL_WD:
        expected_pro = round(annual / PAY_PERIODS_PER_YEAR, 2)
    else:
        expected_pro = round((annual / ANNUAL_WORKING_DAYS) * days_worked, 2)

    reg      = oval(or_, REG_AMT_IDX)
    to_      = oval(or_, TO_AMT_IDX)
    out_base = round(reg + to_, 2)

    if abs(expected_pro - out_base) < VARIANCE_THRESHOLD:
        match_str = 'Match'
        mfill     = GREEN
    else:
        match_str = 'Needs Review'
        mfill     = RED

    ws3.append([
        ir[IN_NAME], ir[IN_EMP_ID], dept_slug(dept), annual,
        join_dt.strftime('%m/%d/%Y'), TOTAL_WD, days_worked,
        expected_pro, reg, to_, out_base, match_str
    ])
    r = ws3.max_row
    ws3.cell(r, 12).fill = mfill
    if r % 2 == 0:
        for c in range(1, 12):
            ws3.cell(r, c).fill = ALT

fmt_currency(ws3, ['D', 'H', 'I', 'J', 'K'])
set_col_widths(ws3, [32, 10, 18, 14, 13, 14, 12, 20, 15, 15, 16, 14])

# ─── SHEET 4: PAY ITEM FLAGS ─────────────────────────────────────────────────
# Source of truth: Pay Items tab — all types compared against Gusto output columns
ws4 = wb.create_sheet('Pay Item Flags')

header_row(ws4, [
    'Employee Name', 'Emp ID', 'Department',
    'Pay Item', 'Status', 'Description', 'Input Amount (Pay Items)', 'Output Amount (Gusto)', 'Difference'
])
freeze(ws4)

def append_pi_flag(ws, name, eid, dept, item_label, status, desc, in_amt, out_amt):
    diff = round(in_amt - out_amt, 2)
    ws.append([name, eid, dept_slug(dept) if dept else '',
               item_label, status, desc, in_amt, out_amt, diff])
    r = ws.max_row
    if status == 'Match':
        flag_fill = GREEN
    elif status == 'Not in Gusto':
        flag_fill = ORANGE
    else:
        flag_fill = RED
    ws.cell(r, 5).fill = flag_fill
    if r % 2 == 0:
        for c in [1, 2, 3, 4, 6, 7, 8, 9]:
            ws.cell(r, c).fill = ALT

# ── Matched employees: compare Pay Items vs Gusto for every pay item type ────
for ir, or_, dept in matched:
    eid  = ir[IN_EMP_ID]
    name = ir[IN_NAME]
    emp_pi = pi_totals.get(eid, {})

    # Collect all canonical types this employee has (from Pay Items or Gusto)
    types_to_check = set(emp_pi.keys())
    # Also include types that Gusto has even if Pay Items doesn't
    for canon, label, out_idx in PAYITEM_GUSTO_MAP:
        if out_idx >= 0 and oval(or_, out_idx) != 0:
            types_to_check.add(canon)

    for canon, label, out_idx in PAYITEM_GUSTO_MAP:
        if canon not in types_to_check:
            continue
        in_amt  = emp_pi.get(canon, 0.0)
        out_amt = oval(or_, out_idx) if out_idx >= 0 else 0.0
        if not in_amt and not out_amt:
            continue
        desc    = pi_notes.get(eid, {}).get(canon, '')
        if abs(in_amt - out_amt) < VARIANCE_THRESHOLD:
            status = 'Match'
        elif out_idx < 0:
            status = 'No Gusto column'
        else:
            status = 'Mismatch'
        append_pi_flag(ws4, name, eid, dept, label, status, desc, in_amt, out_amt)

# ── Unmatched employees: flag any pay items (Gusto = 0) ──────────────────────
for ir, dept, miss_reason in input_only:
    eid     = ir[IN_EMP_ID]
    name    = ir[IN_NAME]
    emp_pi  = pi_totals.get(eid, {})
    for canon, label, out_idx in PAYITEM_GUSTO_MAP:
        in_amt = emp_pi.get(canon, 0.0)
        if in_amt >= VARIANCE_THRESHOLD:
            desc = pi_notes.get(eid, {}).get(canon, '')
            append_pi_flag(ws4, name, eid, dept, label, 'Not in Gusto', desc, in_amt, 0.0)

fmt_currency(ws4, ['G', 'H', 'I'])
set_col_widths(ws4, [32, 10, 18, 18, 14, 36, 22, 22, 12])

# ─── SHEET 5: ZERO / NEGATIVE PAYROLL ────────────────────────────────────────
ws5 = wb.create_sheet('Zero or Negative Payroll')

header_row(ws5, [
    'Last Name', 'First Name', 'Emp ID', 'Department',
    'Regular Amount', 'Net Pay', 'Flag'
])
freeze(ws5)

for or_ in out_rows:
    reg = oval(or_, REG_AMT_IDX)
    net = oval(or_, NET_IDX)
    if reg <= 0 or net <= 0:
        flag = ('Zero Regular' if reg == 0 else
                'Negative Regular' if reg < 0 else
                'Zero Net' if net == 0 else 'Negative Net')
        ws5.append([or_[0], or_[1], get_gusto_emp_id(or_), dept_slug(or_[2]), reg, net, flag])
        r = ws5.max_row
        ws5.cell(r, 7).fill = RED if reg < 0 or net < 0 else WARN
        if r % 2 == 0:
            for c in [1, 2, 3, 4, 5, 6]:
                ws5.cell(r, c).fill = ALT

fmt_currency(ws5, ['E', 'F'])
set_col_widths(ws5, [22, 22, 10, 18, 16, 14, 16])

# ─── SHEET 6: HOURS OVER THRESHOLD ───────────────────────────────────────────
ws_hrs = wb.create_sheet('Hours Over Threshold')
header_row(ws_hrs, [
    'Employee Name', 'Emp ID', 'Department', 'Regular Hrs', 'PTO Hrs',
    'Total Hrs', 'Over By'
])
# No Paid Holidays column in July 16-31 output
freeze(ws_hrs)

RED_FILL = PatternFill('solid', fgColor='FF9999')

for ir, or_, dept in matched:
    reg_hrs = oval(or_, REG_HRS_IDX)
    pto_hrs = oval(or_, TO_HRS_IDX)
    total   = reg_hrs + pto_hrs
    if total > HOURS_THRESHOLD:
        over_by = round(total - HOURS_THRESHOLD, 2)
        ws_hrs.append([
            ir[IN_NAME], ir[IN_EMP_ID], dept_slug(dept),
            reg_hrs, pto_hrs,
            round(total, 2), over_by
        ])
        r = ws_hrs.max_row
        fill = RED_FILL if over_by > 8 else WARN
        for c in range(1, 7):
            ws_hrs.cell(r, c).fill = fill

set_col_widths(ws_hrs, [35, 10, 18, 12, 12, 12, 12])

# ─── SHEET 7: MISSING EMPLOYEES ──────────────────────────────────────────────
ws6 = wb.create_sheet('Missing Employees')
freeze(ws6)

SECTION_FILL_RED    = PatternFill('solid', fgColor='C00000')
SECTION_FILL_ORANGE = PatternFill('solid', fgColor='ED7D31')
SECTION_FILL_GRAY   = PatternFill('solid', fgColor='7F7F7F')
SECTION_FONT        = Font(color='FFFFFF', bold=True, size=11)

def section_header(ws, title, num_cols=6):
    ws.append([title] + [''] * (num_cols - 1))
    r = ws.max_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
    cell = ws.cell(r, 1)
    cell.font      = SECTION_FONT
    cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[r].height = 22
    return r

COL_HEADERS = ['Employee Name', 'Emp ID', 'Department', 'Annual Salary', 'Remote Status', 'Action Needed', 'Notes']

needs_action = [(ir, dept, reason) for ir, dept, reason in input_only
                if reason not in ('LOA',)]

section_header(ws6, '⚠  SHOULD BE PAID — NOT IN GUSTO OUTPUT  (Needs Investigation)', num_cols=7)
ws6[ws6.max_row][0].fill = SECTION_FILL_RED
header_row(ws6, COL_HEADERS)

for ir, dept, miss_reason in needs_action:
    if not ir[IN_NAME].strip():
        continue
    ws6.append([
        ir[IN_NAME], ir[IN_EMP_ID], dept_slug(dept),
        parse_amount(ir[IN_ANNUAL]),
        ir[IN_STATUS],
        'Verify — employee should be paid',
        miss_reason
    ])
    r = ws6.max_row
    ws6.cell(r, 6).fill = RED
    if r % 2 == 0:
        for c in [1, 2, 3, 4, 5, 7]:
            ws6.cell(r, c).fill = ALT

ws6.append([])

loa_only = [(ir, dept, reason) for ir, dept, reason in input_only
            if reason == 'LOA']

section_header(ws6, '✓  ON LEAVE (LOA) — Not expected in Gusto output', num_cols=7)
ws6[ws6.max_row][0].fill = SECTION_FILL_GRAY
header_row(ws6, COL_HEADERS)

for ir, dept, miss_reason in loa_only:
    if not ir[IN_NAME].strip():
        continue
    ws6.append([
        ir[IN_NAME], ir[IN_EMP_ID], dept_slug(dept),
        parse_amount(ir[IN_ANNUAL]),
        ir[IN_STATUS],
        'No action — employee on LOA',
        'LOA'
    ])
    r = ws6.max_row
    ws6.cell(r, 6).fill = BLUE
    if r % 2 == 0:
        for c in [1, 2, 3, 4, 5, 7]:
            ws6.cell(r, c).fill = ALT

ws6.append([])

section_header(ws6, '⚠  IN GUSTO BUT NOT IN REMOTE INPUT  (Verify — may be overpaid)', num_cols=7)
ws6[ws6.max_row][0].fill = SECTION_FILL_ORANGE
header_row(ws6, COL_HEADERS)

for or_ in output_only:
    ws6.append([
        f"{or_[0]}, {or_[1]}", get_gusto_emp_id(or_), dept_slug(or_[2]),
        '', '',
        'Verify — no Remote record found',
        'In Gusto output but not in Remote payroll input'
    ])
    r = ws6.max_row
    ws6.cell(r, 6).fill = WARN
    if r % 2 == 0:
        for c in [1, 2, 3, 4, 5, 7]:
            ws6.cell(r, c).fill = ALT

fmt_currency(ws6, ['D'])
set_col_widths(ws6, [35, 10, 18, 14, 12, 30, 40])

# ─── SHEET 8: ARCHIVED EMPLOYEES ─────────────────────────────────────────────
ws7 = wb.create_sheet('Archived Employees')

header_row(ws7, [
    'Employee Name', 'Emp ID', 'Department', 'Annual Salary',
    'Period Gross (Input)', 'Status', 'Join Date', 'Term Date'
])
freeze(ws7)

for ir, or_, dept in matched:
    if ir[IN_STATUS].strip().lower() == 'archived':
        ws7.append([
            ir[IN_NAME], ir[IN_EMP_ID], dept_slug(dept),
            parse_amount(ir[IN_ANNUAL]),
            parse_amount(ir[IN_PERIOD]),
            ir[IN_STATUS], ir[IN_JOIN], ir[IN_TERM]
        ])
        r = ws7.max_row
        ws7.cell(r, 6).fill = RED
        if r % 2 == 0:
            for c in [1, 2, 3, 4, 5, 7, 8]:
                ws7.cell(r, c).fill = ALT

for ir, dept, _ in input_only:
    if ir[IN_STATUS].strip().lower() == 'archived':
        ws7.append([
            ir[IN_NAME], ir[IN_EMP_ID], dept_slug(dept),
            parse_amount(ir[IN_ANNUAL]),
            parse_amount(ir[IN_PERIOD]),
            ir[IN_STATUS], ir[IN_JOIN], ir[IN_TERM]
        ])
        r = ws7.max_row
        ws7.cell(r, 6).fill = RED

fmt_currency(ws7, ['D', 'E'])
set_col_widths(ws7, [35, 10, 18, 14, 16, 12, 12, 12])

# ─── SHEET: RETRO ────────────────────────────────────────────────────────────
ws_retro = wb.create_sheet('Retro')
header_row(ws_retro, [
    'Employee Name', 'Emp ID', 'Department',
    'Input Retro (Pay Items)', 'Output Retro (Gusto)', 'Difference', 'Match?', 'Note'
])
freeze(ws_retro)

# Build emp_id -> (ir, dept) lookup from matched rows
empid_to_matched = {ir[IN_EMP_ID]: (ir, dept) for ir, or_, dept in matched}

# All retros from Pay Items input
all_retro_emp_ids = set(retro_by_emp_id.keys())
# Also include any output retros not in Pay Items
out_retro_by_empid = {}
for ir, or_, dept in matched:
    out_retro = oval(or_, RETRO_IDX)
    if out_retro:
        out_retro_by_empid[ir[IN_EMP_ID]] = (ir, or_, dept, out_retro)

all_retro_ids = all_retro_emp_ids | set(out_retro_by_empid.keys())

for eid in sorted(all_retro_ids):
    in_info  = retro_by_emp_id.get(eid, {})
    in_amt   = in_info.get('amount', 0.0)
    in_note  = in_info.get('note', '')
    in_name  = in_info.get('name', '')

    if eid in out_retro_by_empid:
        ir, or_, dept, out_amt = out_retro_by_empid[eid]
        emp_name = ir[IN_NAME]
    elif eid in empid_to_matched:
        ir, dept = empid_to_matched[eid]
        out_amt  = 0.0
        emp_name = ir[IN_NAME]
    else:
        emp_name = in_name
        dept     = ''
        out_amt  = 0.0

    diff  = round(in_amt - out_amt, 2)
    match = 'Match' if abs(diff) < VARIANCE_THRESHOLD else 'Needs Review'
    mfill = GREEN if match == 'Match' else RED

    ws_retro.append([emp_name, eid, dept_slug(dept), in_amt, out_amt, diff, match, in_note])
    r = ws_retro.max_row
    ws_retro.cell(r, 7).fill = mfill
    if r % 2 == 0:
        for c in [1, 2, 3, 4, 5, 6, 8]:
            ws_retro.cell(r, c).fill = ALT

fmt_currency(ws_retro, ['D', 'E', 'F'])
set_col_widths(ws_retro, [35, 10, 18, 22, 22, 12, 14, 40])

# ─── SHEET: EXPENSES (Pay Items tab vs Gusto Reimbursements) ─────────────────
# SOURCE OF TRUTH: Pay Items tab (expense types: Expense Non-taxable,
#   Expense Taxable, Mileage Non-taxable, Per diem Non-taxable, Per diem Taxable)
# Compared against Gusto "Reimbursements" output column.
# Shows ALL employees with expenses on either side (matches + mismatches + not in Gusto).
ws_exp = wb.create_sheet('Expenses')
header_row(ws_exp, [
    'Employee Name', 'Emp ID', 'Department',
    'Input Expenses\n(Pay Items)', 'Output Reimbursements\n(Gusto)',
    'Difference', 'Match?', 'Expenses Description'
])
freeze(ws_exp)

def append_exp(ws, name, eid, dept, in_exp, out_exp, note, status_override=None):
    diff  = round(in_exp - out_exp, 2)
    if status_override:
        match = status_override
    else:
        match = 'Match' if abs(diff) < VARIANCE_THRESHOLD else 'Mismatch'
    mfill = GREEN if match == 'Match' else (ORANGE if match == 'Not in Gusto' else RED)
    ws.append([name, eid, dept_slug(dept) if dept else '', in_exp, out_exp, diff, match, note])
    r = ws.max_row
    ws.cell(r, 7).fill = mfill
    if r % 2 == 0:
        for c in [1, 2, 3, 4, 5, 6, 8]:
            ws.cell(r, c).fill = ALT

# Matched employees
for ir, or_, dept in matched:
    eid     = ir[IN_EMP_ID]
    name    = ir[IN_NAME]
    in_exp  = pi_totals.get(eid, {}).get('expense', 0.0)
    out_exp = oval(or_, REIMB_IDX) if REIMB_IDX >= 0 else 0.0
    if in_exp == 0 and out_exp == 0:
        continue
    note = ir[IN_EXP_DESC].strip() if IN_EXP_DESC >= 0 else ''
    append_exp(ws_exp, name, eid, dept, in_exp, out_exp, note)

# Input-only employees with expenses
for ir, dept, miss_reason in input_only:
    eid    = ir[IN_EMP_ID]
    in_exp = pi_totals.get(eid, {}).get('expense', 0.0)
    if in_exp < VARIANCE_THRESHOLD:
        continue
    note = ir[IN_EXP_DESC].strip() if IN_EXP_DESC >= 0 else ''
    append_exp(ws_exp, ir[IN_NAME], eid, dept, in_exp, 0.0, note, status_override='Not in Gusto')

fmt_currency(ws_exp, ['D', 'E', 'F'])
set_col_widths(ws_exp, [35, 10, 18, 20, 20, 12, 14, 45])


# ─── SHEET: GROSS UP (Allowances & Stipends — Payroll Summary vs Gusto) ─────
# Uses Payroll Summary tab as input source (Total Allowance + Total Stipend)
# compared against Gusto Allowance + WFH Stipend output columns.
ws_gross = wb.create_sheet('Gross Up')
header_row(ws_gross, [
    'Employee Name', 'Emp ID', 'Department',
    'Input Allowance\n(Payroll Summary)', 'Input Stipend\n(Payroll Summary)', 'Combined Input',
    'Output Allowance\n(Gusto)', 'Output WFH Stipend\n(Gusto)', 'Combined Output',
    'Difference', 'Match?', 'Allowance Description', 'Stipend Description'
])
freeze(ws_gross)

for ir, or_, dept in matched:
    in_allow = get_amount(ir, IN_ALLOW) if IN_ALLOW >= 0 else 0.0
    in_stip  = get_amount(ir, IN_STIP)  if IN_STIP  >= 0 else 0.0
    combined_in = round(in_allow + in_stip, 2)

    out_allow = oval(or_, ALLOW_OUT_IDX)
    out_stip  = oval(or_, WFH_OUT_IDX)
    combined_out = round(out_allow + out_stip, 2)

    if combined_in == 0 and combined_out == 0:
        continue

    diff  = round(combined_in - combined_out, 2)
    match = 'Match' if abs(diff) < VARIANCE_THRESHOLD else 'Mismatch'
    if match == 'Match':
        continue
    mfill = GREEN if match == 'Match' else RED

    allow_desc = get_desc(ir, IN_ALLOW_DESC)
    stip_desc  = get_desc(ir, IN_STIP_DESC)

    ws_gross.append([
        ir[IN_NAME], ir[IN_EMP_ID], dept_slug(dept),
        in_allow, in_stip, combined_in,
        out_allow, out_stip, combined_out,
        diff, match, allow_desc, stip_desc
    ])
    r = ws_gross.max_row
    ws_gross.cell(r, 11).fill = mfill
    if r % 2 == 0:
        for c in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 12, 13]:
            ws_gross.cell(r, c).fill = ALT

# Also include archived employees from input_only (no Gusto output row — output = 0)
for ir, dept, miss_reason in input_only:
    if ir[IN_STATUS].strip().lower() != 'archived':
        continue
    in_allow = get_amount(ir, IN_ALLOW) if IN_ALLOW >= 0 else 0.0
    in_stip  = get_amount(ir, IN_STIP)  if IN_STIP  >= 0 else 0.0
    combined_in = round(in_allow + in_stip, 2)
    if combined_in == 0:
        continue
    allow_desc = get_desc(ir, IN_ALLOW_DESC)
    stip_desc  = get_desc(ir, IN_STIP_DESC)
    ws_gross.append([
        ir[IN_NAME], ir[IN_EMP_ID], dept_slug(dept) if dept else '',
        in_allow, in_stip, combined_in,
        0.0, 0.0, 0.0,
        combined_in, 'Not in Gusto', allow_desc, stip_desc
    ])
    r = ws_gross.max_row
    ws_gross.cell(r, 11).fill = ORANGE
    if r % 2 == 0:
        for c in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 12, 13]:
            ws_gross.cell(r, c).fill = ALT

fmt_currency(ws_gross, ['D', 'E', 'F', 'G', 'H', 'I', 'J'])
set_col_widths(ws_gross, [35, 10, 18, 18, 16, 16, 18, 18, 16, 12, 12, 35, 35])

# ─── SUMMARY SHEET ───────────────────────────────────────────────────────────
ws0 = wb.create_sheet('Summary', 0)
ws0.sheet_view.showGridLines = False

in_emp_count = len([r for r in in_rows if r and r[IN_NAME].strip()])

summary_data = [
    ('Pay Period',             'July 16 – 31, 2026'),
    ('Pay Date',               'July 31, 2026'),
    ('Working Days',           TOTAL_WD),
    ('Total Input Employees',  in_emp_count),
    ('Total Output Employees', len(out_rows)),
    ('Matched Employees',      len(matched)),
    ('', ''),
    ('Sheet',                  'Count'),
    ('Base Salary Variances',  ws1.max_row - 1),
    ('Term Employees',         ws2.max_row - 1),
    ('New Hire Employees',     ws3.max_row - 1),
    ('Pay Item Flags',         ws4.max_row - 1),
    ('Zero/Negative Payroll',  ws5.max_row - 1),
    ('Hours Over Threshold',   ws_hrs.max_row - 1),
    ('Missing – Needs Action', len(needs_action)),
    ('Missing – LOA',          len(loa_only)),
    ('In Gusto Not Remote',    len(output_only)),
    ('Archived Employees',     ws7.max_row - 1),
    ('Retro Payments',         ws_retro.max_row - 1),
    ('Expenses',               ws_exp.max_row - 1),
    ('Gross Up (Allow/Stip)',  ws_gross.max_row - 1),
]

ws0.column_dimensions['A'].width = 28
ws0.column_dimensions['B'].width = 22

title_font = Font(bold=True, size=14, color='1F4E79')
ws0['A1'] = 'Payroll Audit Report'
ws0['A1'].font = title_font
ws0['A2'] = 'Remote → Gusto  |  July 16–31, 2026'
ws0['A2'].font = Font(size=11, italic=True, color='666666')

for i, (label, value) in enumerate(summary_data, start=4):
    ws0.cell(i, 1).value = label
    ws0.cell(i, 2).value = value
    if label == 'Sheet':
        ws0.cell(i, 1).font = H_FONT
        ws0.cell(i, 1).fill = H_FILL
        ws0.cell(i, 2).font = H_FONT
        ws0.cell(i, 2).fill = H_FILL
    elif label:
        ws0.cell(i, 1).font = Font(bold=True, size=10)

# ─── SAVE ────────────────────────────────────────────────────────────────────
wb.save(RESULT_PATH)
print(f'Saved: {RESULT_PATH}')
print(f'  Working days in period: {TOTAL_WD}')
print(f'  Input employees:        {in_emp_count}')
print(f'  Output employees:       {len(out_rows)}')
print(f'  Matched:                {len(matched)}')
print(f'  Input-only:             {len(input_only)}')
print(f'  Output-only:            {len(output_only)}')
print(f'  Base Salary Variances:  {ws1.max_row - 1}')
print(f'  Term employees:         {ws2.max_row - 1}')
print(f'  New hires:              {ws3.max_row - 1}')
print(f'  Pay Item Flags:         {ws4.max_row - 1}')
print(f'  Zero/Negative payroll:  {ws5.max_row - 1}')
print(f'  Hours Over Threshold:   {ws_hrs.max_row - 1}')
print(f'  Missing – needs action: {len(needs_action)}')
print(f'  Missing – LOA:          {len(loa_only)}')
print(f'  In Gusto not Remote:    {len(output_only)}')
print(f'  Archived:               {ws7.max_row - 1}')
